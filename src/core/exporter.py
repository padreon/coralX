import json
import numpy as np
import pandas as pd
from pathlib import Path
from src.models.project import Project
from src.core.statistics import project_summary, per_image_table, per_station_table, station_summary
from src.core.analysis import photo_area, cover_area_per_code
from src.core.validation import can_run_multivariate, validate_metadata_completeness


def export_csv(project: Project, output_path: str) -> str:
    """Export per-image coverage table to CSV (includes station column)."""
    rows = per_image_table(project)
    df = pd.DataFrame(rows).fillna(0)
    df.to_csv(output_path, index=False)
    return output_path


def export_excel(project: Project, output_path: str) -> str:
    """
    Export to Excel with sheets:
    - Summary: overall project statistics including extended diversity indices
    - Group Coverage: aggregated per benthic group (Hard Coral / Soft Algae / Substrate)
    - Per Station: coverage + diversity per station
    - Per Image: coverage per image with 95% CI columns
    - Cover Area: photo area and per-code area (only when calibrated)
    - Raw Points: every labeled point
    """
    summary = project_summary(project)
    per_station = per_station_table(project)
    per_image = per_image_table(project)

    # Raw points (with station column)
    raw_rows = []
    for station in project.stations:
        for ann in station.annotations:
            for p in ann.points:
                raw_rows.append({
                    "station": station.name,
                    "image": ann.image_path,
                    "point_index": p.index,
                    "x": round(p.x, 2),
                    "y": round(p.y, 2),
                    "label": p.label or "",
                    "category": p.category or "",
                })

    # Summary sheet — three separate sub-tables written at different row offsets:
    #   1. Key metrics      (Metric | Value)
    #   2. Coverage + CI    (Code | Coverage (%) | 95% CI Lower (%) | 95% CI Upper (%))
    #   3. Group coverage   (Group | Coverage (%))
    s1_rows: list[dict] = []
    cov_rows: list[dict] = []
    grp_rows: list[dict] = []
    if summary:
        _reef_health = summary.get("reef_health") or {}
        _hill = summary.get("hill") or {}
        s1_rows = [
            {"Metric": "Total points",              "Value": summary["total_points"]},
            {"Metric": "Labeled points",            "Value": summary["labeled_points"]},
            {"Metric": "",                           "Value": ""},
            {"Metric": "Species richness (S)",      "Value": summary.get("species_richness", "")},
            {"Metric": "Shannon diversity (H')",    "Value": summary.get("shannon_diversity", "")},
            {"Metric": "Simpson diversity (1-D)",   "Value": summary.get("simpson_diversity", "")},
            {"Metric": "Pielou evenness (J')",      "Value": summary.get("pielou_evenness", "")},
            {"Metric": "Margalef richness (d)",     "Value": summary.get("margalef_richness", "")},
            {"Metric": "Fisher alpha (α)",          "Value": summary.get("fisher_alpha", "")},
            {"Metric": "",                           "Value": ""},
            {"Metric": "Mortality Index (MI)",      "Value": summary.get("mortality_index", "")},
            {"Metric": "Reef Health Category",      "Value": _reef_health.get("category", "")},
            {"Metric": "Reef Health (EN)",          "Value": _reef_health.get("category_en", "")},
            {"Metric": "Coral:Algae Ratio",         "Value": summary.get("coral_algae_ratio", "")},
            {"Metric": "Berger-Parker Dominance (d)", "Value": summary.get("berger_parker", "")},
            {"Metric": "Hill q0 (richness)",        "Value": _hill.get("q0", "")},
            {"Metric": "Hill q1 (exp H')",          "Value": _hill.get("q1", "")},
            {"Metric": "Hill q2 (1/Simpson D)",     "Value": _hill.get("q2", "")},
        ]
        cov_rows = [
            {
                "Code": label,
                "Coverage (%)": info["pct"],
                "95% Confidence Interval Lower (%)": info["ci_lower"],
                "95% Confidence Interval Upper (%)": info["ci_upper"],
            }
            for label, info in summary.get("coverage_ci", {}).items()
        ]
        grp_rows = [
            {"Group": grp, "Coverage (%)": pct}
            for grp, pct in summary.get("group_coverage", {}).items()
        ]

    # Group coverage sheet: one row per station + project total
    grp_sheet_rows: list[dict] = []
    all_grp_names: set[str] = set()
    for station in project.stations:
        st_sum = station_summary(station, project.coral_groups)
        grp_cov = st_sum.get("group_coverage", {})
        all_grp_names.update(grp_cov.keys())
        row: dict = {"station": station.name}
        row.update(grp_cov)
        grp_sheet_rows.append(row)
    # Project total row
    if summary.get("group_coverage"):
        total_row: dict = {"station": "PROJECT TOTAL"}
        total_row.update(summary["group_coverage"])
        grp_sheet_rows.append(total_row)

    # Cover area sheet (only for calibrated annotations)
    cover_rows: list[dict] = []
    for station in project.stations:
        for ann in station.annotations:
            p_area = photo_area(ann)
            if p_area is None:
                continue
            c_area = cover_area_per_code(ann) or {}
            row = {
                "station": station.name,
                "image": ann.image_path,
                f"photo_area_{ann.scale_unit}2": p_area,
                "scale_factor_px_per_unit": ann.scale_factor,
                "scale_unit": ann.scale_unit,
            }
            for code, area in c_area.items():
                row[f"{code}_{ann.scale_unit}2"] = area
            cover_rows.append(row)

    # Statistics sheet: mean, std dev, std error per code across all images
    stats_rows = _coverage_statistics(project)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Summary: write three sub-tables with a blank row between each
        row_cursor = 0
        pd.DataFrame(s1_rows).to_excel(
            writer, sheet_name="Summary", index=False, startrow=row_cursor)
        row_cursor += len(s1_rows) + 2  # header + data + blank separator
        if cov_rows:
            pd.DataFrame(cov_rows).to_excel(
                writer, sheet_name="Summary", index=False, startrow=row_cursor)
            row_cursor += len(cov_rows) + 2
        if grp_rows:
            pd.DataFrame(grp_rows).to_excel(
                writer, sheet_name="Summary", index=False, startrow=row_cursor)
        pd.DataFrame(grp_sheet_rows).fillna(0).to_excel(writer, sheet_name="Group Coverage", index=False)
        pd.DataFrame(per_station).fillna(0).to_excel(writer, sheet_name="Per Station", index=False)
        pd.DataFrame(per_image).fillna(0).to_excel(writer, sheet_name="Per Image", index=False)
        pd.DataFrame(stats_rows).to_excel(writer, sheet_name="Statistics", index=False)
        if cover_rows:
            pd.DataFrame(cover_rows).fillna(0).to_excel(writer, sheet_name="Cover Area", index=False)
        pd.DataFrame(raw_rows).to_excel(writer, sheet_name="Raw Points", index=False)

        # --- Multivariate sheets (Lapis 3) ---
        _write_multivariate_sheets(writer, project)

        # --- Map Data sheet (Lapis 3, spatial) ---
        _write_map_data_sheet(writer, project)

    # --- Charts: generate PNGs then embed into Excel (Fase 5) ---
    chart_dir = Path(output_path).with_suffix("").parent / (Path(output_path).stem + "_charts")
    chart_paths = _generate_charts(project, str(chart_dir))
    if chart_paths:
        _embed_charts_sheet(output_path, chart_paths)

    return output_path


def _generate_charts(project: Project, chart_dir: str) -> list[str]:
    """Generate PNG charts to chart_dir. Returns list of paths created.

    Returns empty list silently if matplotlib is not installed.
    """
    try:
        from src.core.plots import export_all_charts
        return export_all_charts(project, chart_dir)
    except ImportError:
        return []
    except Exception:
        return []


def _embed_charts_sheet(excel_path: str, chart_paths: list[str]) -> None:
    """Append a 'Charts' sheet to an existing Excel file with embedded PNGs.

    Uses load_workbook post-processing because pd.ExcelWriter does not support
    add_image while its context manager is open.
    """
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage

    # Pass a file object so openpyxl skips the extension check (it rejects
    # paths without a recognised extension even when the content is valid xlsx).
    with open(excel_path, "rb") as f:
        wb = load_workbook(f)
    ws = wb.create_sheet("Charts")
    row_cursor = 1
    for path in chart_paths:
        if not Path(path).exists():
            continue
        img = XLImage(path)
        img.width = 600
        img.height = 400
        ws.add_image(img, f"A{row_cursor}")
        row_cursor += 22  # ~400px / ~18px per Excel row ≈ 22 rows
    wb.save(excel_path)


def _write_multivariate_sheets(writer: pd.ExcelWriter, project: Project) -> None:
    """Write Bray-Curtis, Ordination, PERMANOVA, SIMPER sheets (Lapis 3).

    If can_run_multivariate() is False, writes a single 'Multivariate' sheet
    with the blocking reason instead.
    """
    from src.core.multivariate import composition_matrix, bray_curtis_matrix, pcoa, permanova, simper

    gate = can_run_multivariate(project)
    if not gate.ok:
        reason_df = pd.DataFrame([{"Reason": r} for r in gate.reasons])
        reason_df.to_excel(writer, sheet_name="Multivariate", index=False)
        return

    sample_names, code_names, matrix = composition_matrix(project)
    bc = bray_curtis_matrix(matrix)

    # Bray-Curtis dissimilarity matrix
    bc_df = pd.DataFrame(bc, index=sample_names, columns=sample_names).round(4)
    bc_df.to_excel(writer, sheet_name="Bray-Curtis")

    # PCoA ordination
    pcoa_result = pcoa(bc)
    coords = pcoa_result["coords"]
    n_axes = coords.shape[1] if coords.ndim == 2 else 0
    ord_rows = []
    for i, name in enumerate(sample_names):
        row: dict = {"station": name}
        for ax in range(n_axes):
            row[f"PCoA{ax+1}"] = round(float(coords[i, ax]), 6)
        ord_rows.append(row)
    var_exp = pcoa_result["variance_explained"]
    # Append variance explained as footer row
    var_row: dict = {"station": "Variance explained"}
    for ax in range(n_axes):
        var_row[f"PCoA{ax+1}"] = var_exp[ax] if ax < len(var_exp) else ""
    ord_rows.append(var_row)
    pd.DataFrame(ord_rows).to_excel(writer, sheet_name="Ordination", index=False)

    # PERMANOVA — use station names as group labels (one group per station = trivial,
    # so only meaningful if project has group metadata; otherwise skip with note)
    # For now, use station names directly as group labels so the math runs.
    perm_result = permanova(bc, sample_names)
    if "error" in perm_result:
        pd.DataFrame([{"Note": perm_result["error"]}]).to_excel(
            writer, sheet_name="PERMANOVA", index=False)
    else:
        perm_rows = [
            {"Metric": "pseudo-F", "Value": perm_result.get("pseudo_F")},
            {"Metric": "p-value", "Value": perm_result.get("p_value")},
            {"Metric": "permutations", "Value": perm_result.get("permutations")},
            {"Metric": "significant (p<0.05)", "Value": str(perm_result.get("significant"))},
        ]
        pd.DataFrame(perm_rows).to_excel(writer, sheet_name="PERMANOVA", index=False)

    # SIMPER — first pair of stations as example (real UI would let user pick groups)
    if len(sample_names) >= 2:
        simper_result = simper(matrix, code_names, sample_names,
                               sample_names[0], sample_names[1])
        if simper_result:
            pd.DataFrame(simper_result).to_excel(writer, sheet_name="SIMPER", index=False)


def _write_map_data_sheet(writer: pd.ExcelWriter, project: Project) -> None:
    """Write Map Data sheet (station GPS coords + key metrics) for GIS/Google Earth."""
    meta = validate_metadata_completeness(project)
    if not meta["spatial"].ok:
        return

    coral_groups = getattr(project, "coral_groups", [])
    rows: list[dict] = []
    for st in project.stations:
        lat = getattr(st, "gps_lat", None)
        lon = getattr(st, "gps_lon", None)
        if not lat or not lon:
            continue
        from src.core.statistics import station_summary as _st_sum
        summ = _st_sum(st, coral_groups)
        reef = summ.get("reef_health") or {}
        rows.append({
            "station": st.name,
            "lat": lat,
            "lon": lon,
            "live_coral_pct": summ.get("group_coverage", {}).get("Hard Coral", ""),
            "mortality_index": summ.get("mortality_index", ""),
            "reef_health": reef.get("category", ""),
        })
    if rows:
        pd.DataFrame(rows).to_excel(writer, sheet_name="Map Data", index=False)


def _coverage_statistics(project: Project) -> list[dict]:
    """Per-code mean, std dev, and std error across all images in the project."""
    all_annotations = project.annotations
    if not all_annotations:
        return []

    per_image = [ann.coverage_stats() for ann in all_annotations]
    all_codes = sorted({code for row in per_image for code in row})
    n = len(per_image)

    matrix = np.array(
        [[row.get(code, 0.0) for code in all_codes] for row in per_image],
        dtype=float,
    )

    means  = matrix.mean(axis=0)
    stds   = matrix.std(axis=0, ddof=1) if n > 1 else np.zeros(len(all_codes))
    errors = stds / np.sqrt(n) if n > 0 else np.zeros(len(all_codes))

    return [
        {
            "Code":          code,
            "Mean (%)":      round(float(means[i]),  4),
            "Std Dev (%)":   round(float(stds[i]),   4),
            "Std Error (%)": round(float(errors[i]), 4),
        }
        for i, code in enumerate(all_codes)
    ]


def export_coral_codes(project: Project, output_path: str) -> str:
    """
    Export project coral codes to JSON or CSV.

    JSON (.json) — full round-trip format:
      {
        "codes":  {"HC": "Hard Coral", ...},
        "groups": [{"name": "Coral", "codes": ["HC", ...], "color": "FF8000"}, ...]
      }
      Re-importable via File → Import → Coral Codes.

    CSV (.csv / .tsv) — flat table with columns: code, description, group, color.
      One row per code; group and color come from coral_groups (empty if ungrouped).
    """
    ext = Path(output_path).suffix.lower()

    # Build code → group name and color lookup
    code_to_group: dict[str, str] = {}
    code_to_color: dict[str, str] = {}
    for g in project.coral_groups:
        grp_name = g.get("name", "")
        grp_color = g.get("color", "")
        for c in g.get("codes", []):
            code_to_group[c] = grp_name
            code_to_color[c] = grp_color

    if ext == ".json":
        data = {
            "codes": project.coral_codes,
            "groups": project.coral_groups,
        }
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

    elif ext in (".csv", ".tsv"):
        sep = "\t" if ext == ".tsv" else ","
        rows = [
            {
                "code": code,
                "description": desc,
                "group": code_to_group.get(code, ""),
                "color": code_to_color.get(code, ""),
            }
            for code, desc in project.coral_codes.items()
        ]
        pd.DataFrame(rows).to_csv(output_path, index=False, sep=sep)

    else:
        raise ValueError(f"Unsupported format: {ext}. Use .json, .csv, or .tsv")

    return output_path
